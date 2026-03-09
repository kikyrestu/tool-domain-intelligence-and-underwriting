"""Export Service — generate CSV/XLSX from candidate data."""

import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.candidate import CandidateDomain

EXPORT_COLUMNS = [
    ("domain", "Domain"),
    ("niche", "Niche"),
    ("availability_status", "Availability"),
    ("is_domain_alive", "Domain Status"),
    ("score_total", "Score"),
    ("label", "Label"),
    ("label_reason", "Reason"),
    ("whois_registrar", "Registrar"),
    ("whois_created_date", "Created"),
    ("whois_expiry_date", "Expires"),
    ("whois_days_left", "Days Left"),
    ("dns_has_records", "DNS"),
    ("wayback_total_snapshots", "Snapshots"),
    ("wayback_first_seen", "First Seen"),
    ("wayback_last_seen", "Last Seen"),
    ("dominant_language", "Language"),
    ("source_url_found", "Source URL"),
    ("owner_notes", "Notes"),
    ("created_at", "Discovered"),
]


def _candidate_row(c: CandidateDomain) -> list:
    """Extract a row of values from a candidate."""
    return [
        c.domain,
        c.niche,
        c.availability_status or "",
        "dead" if not c.is_domain_alive else "alive",
        c.score_total if c.score_total is not None else "",
        c.label or "",
        c.label_reason or "",
        c.whois_registrar or "",
        str(c.whois_created_date) if c.whois_created_date else "",
        str(c.whois_expiry_date) if c.whois_expiry_date else "",
        c.whois_days_left if c.whois_days_left is not None else "",
        "yes" if c.dns_has_records else ("no" if c.dns_has_records is not None else ""),
        c.wayback_total_snapshots or 0,
        str(c.wayback_first_seen) if c.wayback_first_seen else "",
        str(c.wayback_last_seen) if c.wayback_last_seen else "",
        c.dominant_language or "",
        c.source_url_found or "",
        c.owner_notes or "",
        c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "",
    ]


async def _get_candidates(
    db: AsyncSession,
    status_filter: str | None = None,
    niche_filter: str | None = None,
    label_filter: str | None = None,
) -> list[CandidateDomain]:
    query = select(CandidateDomain)
    if status_filter:
        query = query.where(CandidateDomain.availability_status == status_filter)
    if niche_filter:
        query = query.where(CandidateDomain.niche == niche_filter)
    if label_filter:
        query = query.where(CandidateDomain.label == label_filter)
    query = query.order_by(CandidateDomain.score_total.desc().nulls_last())
    result = await db.execute(query)
    return result.scalars().all()


async def generate_csv(
    db: AsyncSession,
    status_filter: str | None = None,
    niche_filter: str | None = None,
    label_filter: str | None = None,
) -> str:
    """Generate CSV string from filtered candidates."""
    candidates = await _get_candidates(db, status_filter, niche_filter, label_filter)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([col[1] for col in EXPORT_COLUMNS])

    for c in candidates:
        writer.writerow(_candidate_row(c))

    return output.getvalue()


# Label → fill color mapping for XLSX
_LABEL_FILLS = {
    "Available": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Watchlist": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "Uncertain": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Discard": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


async def generate_xlsx(
    db: AsyncSession,
    status_filter: str | None = None,
    niche_filter: str | None = None,
    label_filter: str | None = None,
) -> bytes:
    """Generate formatted XLSX bytes from filtered candidates."""
    candidates = await _get_candidates(db, status_filter, niche_filter, label_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = "Domain Candidates"

    # Header row
    headers = [col[1] for col in EXPORT_COLUMNS]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

    # Data rows
    label_col_idx = next(i for i, (k, _) in enumerate(EXPORT_COLUMNS) if k == "label") + 1

    for c in candidates:
        row_data = _candidate_row(c)
        ws.append(row_data)
        row_num = ws.max_row

        # Apply label fill to entire row
        label_fill = _LABEL_FILLS.get(c.label)
        if label_fill:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_idx).fill = label_fill

        # Borders
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_idx).border = _THIN_BORDER

    # Auto-width columns
    for col_idx, (key, header) in enumerate(EXPORT_COLUMNS, 1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 40))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
